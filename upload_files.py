"""
One-time setup script: Create a Gemini File Search Store and upload all 40 GO PDFs with metadata.

Usage:
    1. Set GEMINI_API_KEY environment variable
    2. Run: python upload_files.py

This creates a File Search Store named "TNega-GOs", uploads all PDFs from tnega/ folder,
attaches custom metadata from GO_metadata.xlsx, and saves the store name to store_config.json.
"""

import json
import os
import time

import openpyxl
from dotenv import load_dotenv
from google import genai

load_dotenv()

TNEGA_DIR = "tnega"
METADATA_FILE = "GO_metadata.xlsx"
STORE_CONFIG_FILE = "store_config.json"
STORE_DISPLAY_NAME = "TNega-GOs"
POLL_INTERVAL = 5  # seconds between polling for upload completion


def load_metadata(xlsx_path: str) -> dict[str, dict]:
    """Load GO metadata from Excel. Returns dict keyed by filename."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    metadata_by_file = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        filename = record["Filename"]
        metadata_by_file[filename] = record

    wb.close()
    return metadata_by_file


def build_custom_metadata(record: dict) -> list[dict]:
    """Build the custom_metadata list for a single file from its Excel row."""
    metadata = []

    if record.get("Year") is not None:
        metadata.append({"key": "year", "numeric_value": int(record["Year"])})

    if record.get("GO Number"):
        metadata.append({"key": "go_number", "string_value": str(record["GO Number"])})

    if record.get("Department"):
        metadata.append({"key": "department", "string_value": str(record["Department"])})

    if record.get("Abstract"):
        abstract = str(record["Abstract"])
        # Gemini enforces 256 byte limit; truncate by bytes to be safe
        encoded = abstract.encode("utf-8")[:256]
        abstract = encoded.decode("utf-8", errors="ignore").rstrip()
        metadata.append({"key": "abstract", "string_value": abstract})

    if record.get("Date"):
        metadata.append({"key": "date", "string_value": str(record["Date"])})

    return metadata


def main():
    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        print("ERROR: Set GEMINI_API_KEY environment variable first.")
        return

    client = genai.Client(api_key=api_key)

    # Load metadata from Excel
    print(f"Loading metadata from {METADATA_FILE}...")
    metadata_by_file = load_metadata(METADATA_FILE)
    print(f"  Found metadata for {len(metadata_by_file)} files.")

    # Create File Search Store
    print(f"\nCreating File Search Store '{STORE_DISPLAY_NAME}'...")
    store = client.file_search_stores.create(
        config={"display_name": STORE_DISPLAY_NAME}
    )
    store_name = store.name
    print(f"  Store created: {store_name}")

    # Get list of PDFs to upload
    pdf_files = sorted(f for f in os.listdir(TNEGA_DIR) if f.lower().endswith(".pdf"))
    print(f"\nFound {len(pdf_files)} PDFs in {TNEGA_DIR}/")

    # Upload each PDF with metadata
    operations = []
    for i, pdf_filename in enumerate(pdf_files, 1):
        pdf_path = os.path.join(TNEGA_DIR, pdf_filename)
        record = metadata_by_file.get(pdf_filename, {})
        custom_metadata = build_custom_metadata(record) if record else []

        upload_config = {"display_name": pdf_filename}
        if custom_metadata:
            upload_config["custom_metadata"] = custom_metadata

        print(f"  [{i}/{len(pdf_files)}] Uploading {pdf_filename}...", end=" ", flush=True)

        try:
            operation = client.file_search_stores.upload_to_file_search_store(
                file=pdf_path,
                file_search_store_name=store_name,
                config=upload_config,
            )
            operations.append((pdf_filename, operation))
            print("started.")
        except Exception as e:
            print(f"FAILED: {e}")

    # Wait for all uploads to complete
    print(f"\nWaiting for {len(operations)} uploads to complete...")
    completed = 0
    failed = 0

    for pdf_filename, operation in operations:
        retries = 0
        while not operation.done:
            time.sleep(POLL_INTERVAL)
            try:
                operation = client.operations.get(operation)
            except Exception as e:
                retries += 1
                if retries > 5:
                    print(f"  POLL FAILED: {pdf_filename} - {e}")
                    break
                time.sleep(POLL_INTERVAL)
                continue

        if operation.error:
            print(f"  FAILED: {pdf_filename} - {operation.error}")
            failed += 1
        else:
            completed += 1
            print(f"  OK: {pdf_filename}")

    print(f"\nUpload complete: {completed} succeeded, {failed} failed out of {len(operations)} total.")

    # Save store config for app.py
    config = {"store_name": store_name, "display_name": STORE_DISPLAY_NAME}
    with open(STORE_CONFIG_FILE, "w") as f:
        json.dump(config, f, indent=2)
    print(f"\nStore config saved to {STORE_CONFIG_FILE}")
    print("You can now run: chainlit run app.py")


if __name__ == "__main__":
    main()
